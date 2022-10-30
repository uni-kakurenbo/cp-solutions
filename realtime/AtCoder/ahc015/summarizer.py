import openpyxl
from sys import argv


file = "./summaries/solver26.0.xlsx"
wb = openpyxl.load_workbook(file)

sheet = wb['Score']

L = 500


directory = "cases" + str(L)

for row in sheet.iter_rows(min_row=2, max_row=17, min_col=6, max_col=sheet.max_column):
    for cell in row:
        cell.value = None

for n in range(31, 62, 2):
    sheet.cell((n-29)//2+1,1).value = n;

cnt = [5] * 31;

for i in range(L):
    case_id = str(i).zfill(4)
    case = open(f"{directory}/{case_id}.txt").read()
    score = int(open(f"{directory}/{case_id}.score").read())
    time = int(float(open(f"{directory}/{case_id}.info").read().splitlines()[1]))
    answer_length = int(open(f"{directory}/{case_id}.res").read().splitlines()[0].split()[0])
    n, m = map(int, case.splitlines()[0].split())
    row = (n-29)//2
    cnt[row] += 1
    sheet.cell(row+1, cnt[row]).value = score

wb.save(file)
wb.close()
